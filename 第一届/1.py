import pandas as pd
from datetime import datetime, timedelta
import numpy as np
from openpyxl import load_workbook
# 读取Excel文件的第一个工作表
file_name = "2024年2月考勤明细20241.xlsx"
file_name2 = "24年加班分析-2月.xlsx"
# 请假的日期
leave_date = [7, 8, 9]

# 工作日天数
work_day = 15

# 判断请假0.5天依据小时数
hhhhhh = 7

# 判断加班0.5天依据小时数
hhhhhhhh = 1

json_data = []


# 增加数据
def add_data(name, key, value):
    for item in json_data:
        if name in item:
            item[name].append(key)
            item[name].append(value)
            break
    else:
        json_data.append({name: [key, value]})


# 删除数据
def delete_data(name, key):
    for item in json_data:
        if name in item:
            item[name].remove(key)
            break


# 修改数据
def update_data(name, key, new_value):
    for item in json_data:
        if name in item:
            index = item[name].index(key)
            item[name][index + 1] = new_value
            break


# 查询数据
def query_data(name, key):
    for item in json_data:
        if name in item:
            index = item[name].index(key)
            return item[name][index + 1]
    return None


# 时间差
def time_difference(data):
    """
    计算时间差
    :param data: 时间字符串
    :return: 时间差
    """
    # print(data)
    data = data[:5]
    # 将字符串转换为时间对象
    # print(data)
    time_obj = datetime.strptime(data, '%H:%M')

    # 创建一个表示9.30的时间对象
    target_time = datetime.strptime('09:30', '%H:%M')

    # 比较两个时间对象的大小，如果大于9.30，则计算时间差
    if time_obj > target_time:
        time_diff = time_obj - target_time

        # 返回分钟数
        return time_diff.seconds // 60
    else:
        return "时间未超过9.30"


# 计算考勤调休数
def time_difference_tiao(text):
    # 提取时间部分
    if '补卡申请' in text:
        return 0
    elif '产检假' in text:
        start_time_str = text[9:14]
        end_time_str = text[21:26]
    elif len(text) == 5:
        start_time_str = text[9:14]
        end_time_str = text[21:26]
    else:
        start_time_str = text[8:13]
        end_time_str = text[20:25]
    # print(start_time_str, end_time_str)
    # 将时间字符串转换为datetime对象
    date_format = "%H:%M"
    start_time = datetime.strptime(start_time_str, date_format)
    end_time = datetime.strptime(end_time_str, date_format)
    # print(start_time, end_time)
    # 计算时间差的秒数
    time_diff_seconds = (end_time - start_time).total_seconds()

    # 将秒数转换为小时数
    hours = time_diff_seconds / 3600
    if hours >= hhhhhh:
        return 1
    else:
        return 0.5


def time_difference_jiaban(text,text1):

    start_time_str = text
    end_time_str = text1
    # print(start_time_str, end_time_str)
    # 将时间字符串转换为datetime对象
    date_format = "%H:%M"
    start_time = datetime.strptime(start_time_str, date_format)
    end_time = datetime.strptime(end_time_str, date_format)
    # print(start_time, end_time)
    # 计算时间差的秒数
    time_diff_seconds = (end_time - start_time).total_seconds()

    # 将秒数转换为小时数
    hours = time_diff_seconds / 3600
    if hours >= hhhhhhhh:
        return 1
    else:
        return 0.5

def time_category(time_str):
    hour, minute = map(int, time_str.split(':'))
    hour = hour - 12
    if hour >= 10:
        return "加班到10点以后"
    elif hour >= 9:
        return "加班到9点以后"
    elif hour >= 8:
        return "加班到8点以后"
    elif hour >= 7:
        return "7点以后下班"
    else:
        return "7点以前下班"


names = []

json_shangban = {}
json_xiaban = {}
json_zhongshang = {}
json_zhongxia = {}
df = pd.read_excel(file_name, sheet_name=0, engine='openpyxl')
lists = []
# 从第三行开始输出每一行数据
for index, row in df.iterrows():
    # print(list(row))
    # ['衣健男', '产研中心', '产研中心', nan, '技术总监', '053153452950', '24-02-25 星期日', '1708790400000', '休息',
    #  '09:30', '正常', '14:57', '正常', nan, nan, nan, nan, nan, nan, nan, nan, nan, '1', '1', '327', nan, nan, nan, nan,
    #  nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, '0.5', nan, '0.5', nan]
    #
    name_names = ''
    if index >= 3:
        # print(list(row)[6])
        # 首先，去除字符串中的中文星期部分，以便正确解析日期
        date_only = list(row)[6].split(' ')[0]

        # 然后，根据格式化字符串将字符串转换成datetime对象
        # 注意这里的格式应当对应字符串的实际格式，"%d-%m-%y" 表示日-月-年两位数形式
        date_object = datetime.strptime(date_only, "%y-%m-%d")

        # 从datetime对象中提取日
        day_of_month = date_object.day

        if day_of_month in leave_date:
            if pd.isna(row.iloc[16]):
                # print(str(row.iloc[12]))
                if str(row.iloc[12]) == '正常' or '补卡审批通过' in str(row.iloc[12]) or list(row)[16] == '早退':
                    update_data(name_names, time_category(list(row)[11]),
                                query_data(name_names, time_category(list(row)[11])) + 1)
                    if str(row.iloc[10]) == '正常':
                        update_data(name_names, '打卡', query_data(name_names, '打卡') + 1)
                    if list(row)[8] == '休息':
                        update_data(name_names, '加班', query_data(name_names, '加班') + 1)
                        if query_data(name_names, '非展会周末加班') == '':
                            update_data(name_names, '非展会周末加班', str(day_of_month) + '日加班' + str(
                                time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
                        else:
                            update_data(name_names, '非展会周末加班',
                                        query_data(name_names, '非展会周末加班') + '、' + str(
                                            day_of_month) + '日加班' + str(
                                            time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
                if (list(row)[10] == '正常' or list(row)[10] == '迟到') and (list(row)[16] == '正常' or list(row)[16] == '早退'):
                    update_data(name_names, '打卡', query_data(name_names, '打卡') + 1)
                    if list(row)[6] == '休息':
                        update_data(name_names, '加班', query_data(name_names, '加班') + 1)
                        if query_data(name_names, '非展会周末加班') == '':
                            update_data(name_names, '非展会周末加班', str(day_of_month) + '日加班' + str(
                                time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
                        else:
                            update_data(name_names, '非展会周末加班',
                                        query_data(name_names, '非展会周末加班') + '、' + str(
                                            day_of_month) + '日加班' + str(
                                            time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
            continue
        if list(row)[1] != '未加入考勤组' and list(row)[1] != '运营-会员运营中心-会员营销' or  '离职' in list(row)[0]:
            # print(list(row))
            # 判断姓名是否重复
            if '离职' in list(row)[0]:
                # 替换名字的后缀
                name_names = list(row)[0].replace('（离职）', '')
            else:
                name_names = list(row)[0]
            # print(name_names)
            if name_names not in names:
                names.append(name_names)
                add_data(name_names, '打卡', 0)
                add_data(name_names, '迟到20分钟内', '')
                add_data(name_names, '迟到20分钟以上', '')
                add_data(name_names, '上班未打卡', '')
                add_data(name_names, '下班未打卡', '')
                add_data(name_names, '旷工', '')
                add_data(name_names, '调休', 0)
                add_data(name_names, '出差', '')
                add_data(name_names, '请假', '')
                add_data(name_names, '外出', 0)
                add_data(name_names, '出勤', 0)
                add_data(name_names, '加班到10点以后', 0)
                add_data(name_names, '加班到9点以后', 0)
                add_data(name_names, '加班到8点以后', 0)
                add_data(name_names, '7点以后下班', 0)
                add_data(name_names, '7点以前下班', 0)
                add_data(name_names, '调休请假', 0)
                add_data(name_names, '加班', 0)
                add_data(name_names, '非展会周末加班', '')

                if not pd.isna(row.iloc[21]):  # 检查第21个元素是否为缺失值
                    if '调休' in row.iloc[21]:
                        update_data(name_names, '调休',
                                    query_data(name_names, '调休') + time_difference_tiao(list(row)[21]))
                    if '事假' in row.iloc[21]:
                        if query_data(name_names, '请假') == '':
                            update_data(name_names, '请假', str(day_of_month) + '日' + '事假' + str(
                                time_difference_tiao(list(row)[21])) + '天')
                        else:
                            update_data(name_names, '请假', query_data(name_names, '请假') + ';' + str(
                                day_of_month) + '日' + '事假' + str(time_difference_tiao(list(row)[21])) + '天')

                    if '病假' in row.iloc[21]:
                        if query_data(name_names, '请假') == '':
                            update_data(name_names, '请假', str(day_of_month) + '日' + '病假' + str(
                                time_difference_tiao(list(row)[21])) + '天')
                        else:
                            update_data(name_names, '请假', query_data(name_names, '请假') + ';' + str(
                                day_of_month) + '日' + '病假' + str(time_difference_tiao(list(row)[21])) + '天')
                    if '产假' in row.iloc[21]:
                        if query_data(name_names, '请假') == '':
                            update_data(name_names, '请假', str(day_of_month) + '日' + '产假' + str(
                                time_difference_tiao(list(row)[21])) + '天')
                        else:
                            update_data(name_names, '请假', query_data(name_names, '请假') + ';' + str(
                                day_of_month) + '日' + '产假' + str(time_difference_tiao(list(row)[21])) + '天')
                    if '产检假' in row.iloc[21]:
                        if query_data(name_names, '请假') == '':
                            update_data(name_names, '请假', str(day_of_month) + '日' + '产检假' + str(
                                time_difference_tiao(list(row)[21])) + '天')
                        else:
                            update_data(name_names, '请假', query_data(name_names, '请假') + ';' + str(
                                day_of_month) + '日' + '产检假' + str(time_difference_tiao(list(row)[21])) + '天')
                    if '外出' in row.iloc[21]:
                        if query_data(name_names, '外出') == '':
                            update_data(name_names, '外出',
                                        query_data(name_names, '外出') + time_difference_tiao(list(row)[21]))
                        else:
                            update_data(name_names, '外出',
                                        query_data(name_names, '外出') + time_difference_tiao(list(row)[21]))
                    else:
                        update_data(name_names, '调休请假',
                                    query_data(name_names, '调休请假') + time_difference_tiao(list(row)[21]))
                    # if '调休' in row.iloc[21]:
                    #     update_data(name_names, '调休', query_data(name_names, '调休') + time_difference_tiao(list(row)[21]))
                    # if '事假' in row.iloc[21]:
                    #     update_data(name_names, '旷工', query_data(name_names, '旷工') + time_difference_tiao(list(row)[21]))
                    # if '病假' in row.iloc[21]:
                    #     update_data(name_names, '旷工', query_data(name_names, '旷工') + time_difference_tiao(list(row)[21]))
                    #     update_data(name_names, '出差',   time_difference_tiao(list(row)[21]))
                    # if '病假' in row.iloc[21]:
                    #     update_data(name_names, '旷工', query_data(name_names, '旷工') + time_difference_tiao(list(row)[21]))
                    # if '产假' in row.iloc[21]:
                    #     update_data(name_names, '上班未打卡', query_data(name_names, '上班未打卡') + time_difference_tiao(list(row)[21]))
                    # if '产检假' in row.iloc[21]:
                    #     update_data(name_names, '下班未打卡', query_data(name_names, '下班未打卡') + time_difference_tiao(list(row)[21]))

                if (list(row)[10] == '正常' or list(row)[10] == '迟到') and (list(row)[16] == '正常' or list(row)[16] == '早退'):
                    update_data(name_names, '打卡', query_data(name_names, '打卡') + 1)
                    if list(row)[6] == '休息':
                        update_data(name_names, '加班', query_data(name_names, '加班') + 1)
                        if query_data(name_names, '非展会周末加班') == '':
                            update_data(name_names, '非展会周末加班', str(day_of_month) + '日加班' + str(
                                time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
                        else:
                            update_data(name_names, '非展会周末加班',
                                        query_data(name_names, '非展会周末加班') + '、' + str(
                                            day_of_month) + '日加班' + str(
                                            time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
                if list(row)[10] == '迟到':
                    # print(list(row)[9])
                    tinm_min = time_difference(list(row)[9])
                    if int(tinm_min) <= 20:
                        if query_data(name_names, '迟到20分钟内') == '':
                            update_data(name_names, '迟到20分钟内',
                                        query_data(name_names, '迟到20分钟内') + str(day_of_month) + '日' + str(
                                            tinm_min) + '分')
                        else:
                            update_data(name_names, '迟到20分钟内',
                                        query_data(name_names, '迟到20分钟内') + '、' + str(day_of_month) + '日' + str(
                                            tinm_min) + '分')
                    else:
                        if query_data(name_names, '迟到20分钟以上') == '':
                            update_data(name_names, '迟到20分钟以上',
                                        query_data(name_names, '迟到20分钟以上') + str(day_of_month) + '日' + str(
                                            tinm_min) + '分')
                        else:
                            update_data(name_names, '迟到20分钟以上',
                                        query_data(name_names, '迟到20分钟以上') + '、' + str(
                                            day_of_month) + '日' + str(
                                            tinm_min) + '分')

                if list(row)[10] == '缺卡' and list(row)[12] == '缺卡' and list(row)[14] == '缺卡' and list(row)[
                    16] == '缺卡':

                    if query_data(name_names, '旷工') == '':
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + str(day_of_month) + '日')
                    else:
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + '、' + str(day_of_month) + '日')

                    continue
                if list(row)[10] == '缺卡':
                    if query_data(name_names, '上班未打卡') == '':
                        update_data(name_names, '上班未打卡',
                                    query_data(name_names, '上班未打卡') + str(day_of_month) + '日')

                    else:
                        update_data(name_names, '上班未打卡',
                                    query_data(name_names, '上班未打卡') + '、' + str(day_of_month) + '日')
                    if query_data(name_names, '旷工') == '':
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + str(day_of_month) + '日上午')
                    else:
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + '、' + str(day_of_month) + '日上午')
                if list(row)[16] == '缺卡':

                    if query_data(name_names, '下班未打卡') == '':
                        update_data(name_names, '下班未打卡', str(day_of_month) + '日')
                    else:
                        update_data(name_names, '下班未打卡',
                                    query_data(name_names, '下班未打卡') + '、' + str(day_of_month) + '日')
                    if query_data(name_names, '旷工') == '':
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + str(day_of_month) + '日下午')
                    else:
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + '、' + str(day_of_month) + '日下午')
                if list(row)[16] == '正常' or list(row)[16] == '早退':
                    update_data(name_names,time_category(list(row)[15]),query_data(name_names, time_category(list(row)[15]))+1)
                if pd.isna(row.iloc[16]):
                    # print(str(row.iloc[12]))
                    if str(row.iloc[12]) == '正常' or '补卡审批通过' in str(row.iloc[12]) or list(row)[16] == '早退':
                        update_data(name_names, time_category(list(row)[11]),
                                    query_data(name_names, time_category(list(row)[11])) + 1)
                        if str(row.iloc[10]) == '正常':
                            update_data(name_names, '打卡', query_data(name_names, '打卡') + 1)
                        if list(row)[8] == '休息':
                            update_data(name_names, '加班', query_data(name_names, '加班') + 1)
                            if query_data(name_names, '非展会周末加班') == '':
                                update_data(name_names, '非展会周末加班', str(day_of_month) + '日加班'+ str(time_difference_jiaban(list(row)[9],list(row)[11])) + '天')
                            else:
                                update_data(name_names, '非展会周末加班',
                                            query_data(name_names, '非展会周末加班') + '、' + str(day_of_month) + '日加班'+ str(time_difference_jiaban(list(row)[9],list(row)[11])) + '天')

            else:
                if not pd.isna(row.iloc[21]):  # 检查第21个元素是否为缺失值
                    if '调休' in row.iloc[21]:
                        update_data(name_names, '调休',
                                    query_data(name_names, '调休') + time_difference_tiao(list(row)[21]))
                    if '事假' in row.iloc[21]:
                        if query_data(name_names, '请假') == '':
                            update_data(name_names, '请假', str(day_of_month) + '日' + '事假' + str(
                                time_difference_tiao(list(row)[21])) + '天')
                        else:
                            update_data(name_names, '请假', query_data(name_names, '请假') + ';' + str(
                                day_of_month) + '日' + '事假' + str(time_difference_tiao(list(row)[21])) + '天')
                    if '病假' in row.iloc[21]:
                        if query_data(name_names, '请假') == '':
                            update_data(name_names, '请假', str(day_of_month) + '日' + '病假' + str(
                                time_difference_tiao(list(row)[21])) + '天')
                        else:
                            update_data(name_names, '请假', query_data(name_names, '请假') + ';' + str(
                                day_of_month) + '日' + '病假' + str(time_difference_tiao(list(row)[21])) + '天')
                    if '产假' in row.iloc[21]:
                        if query_data(name_names, '请假') == '':
                            update_data(name_names, '请假', str(day_of_month) + '日' + '产假' + str(
                                time_difference_tiao(list(row)[21])) + '天')
                        else:
                            update_data(name_names, '请假', query_data(name_names, '请假') + ';' + str(
                                day_of_month) + '日' + '产假' + str(time_difference_tiao(list(row)[21])) + '天')
                    if '产检假' in row.iloc[21]:
                        if query_data(name_names, '请假') == '':
                            update_data(name_names, '请假', str(day_of_month) + '日' + '产检假' + str(
                                time_difference_tiao(list(row)[21])) + '天')
                        else:
                            update_data(name_names, '请假', query_data(name_names, '请假') + ';' + str(
                                day_of_month) + '日' + '产检假' + str(time_difference_tiao(list(row)[21])) + '天')
                    if '外出' in row.iloc[21]:
                        if query_data(name_names, '外出') == '':
                            update_data(name_names, '外出',
                                        query_data(name_names, '外出') + time_difference_tiao(list(row)[21]))
                        else:
                            update_data(name_names, '外出',
                                        query_data(name_names, '外出') + time_difference_tiao(list(row)[21]))
                    else:
                        # print(list(row)[21])
                        update_data(name_names, '调休请假',
                                    query_data(name_names, '调休请假') + time_difference_tiao(list(row)[21]))

                if (list(row)[10] == '正常' or list(row)[10] == '迟到') and (list(row)[16] == '正常' or list(row)[16] == '早退'):
                    update_data(name_names, '打卡', query_data(name_names, '打卡') + 1)
                    if list(row)[6] == '休息':
                        update_data(name_names, '加班', query_data(name_names, '加班') + 1)
                        if query_data(name_names, '非展会周末加班') == '':
                            update_data(name_names, '非展会周末加班', str(day_of_month) + '日加班' + str(
                                time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
                        else:
                            update_data(name_names, '非展会周末加班',
                                        query_data(name_names, '非展会周末加班') + '、' + str(
                                            day_of_month) + '日加班' + str(
                                            time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
                if list(row)[10] == '迟到':
                    # print(list(row)[9])
                    tinm_min = time_difference(list(row)[9])
                    # print(tinm_min)
                    if int(tinm_min) <= 20:
                        if query_data(name_names, '迟到20分钟内') == '':
                            update_data(name_names, '迟到20分钟内',
                                        query_data(name_names, '迟到20分钟内') + str(day_of_month) + '日' + str(
                                            tinm_min) + '分')
                        else:
                            update_data(name_names, '迟到20分钟内',
                                        query_data(name_names, '迟到20分钟内') + '、' + str(day_of_month) + '日' + str(
                                            tinm_min) + '分')
                    else:
                        if query_data(name_names, '迟到20分钟以上') == '':
                            update_data(name_names, '迟到20分钟以上',
                                        query_data(name_names, '迟到20分钟以上') + str(day_of_month) + '日' + str(
                                            tinm_min) + '分')
                        else:
                            update_data(name_names, '迟到20分钟以上',
                                        query_data(name_names, '迟到20分钟以上') + '、' + str(
                                            day_of_month) + '日' + str(
                                            tinm_min) + '分')
                if list(row)[10] == '缺卡' and list(row)[12] == '缺卡' and list(row)[14] == '缺卡' and list(row)[
                    16] == '缺卡':
                    if query_data(name_names, '旷工') == '':
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + str(day_of_month) + '日')
                    else:
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + '、' + str(day_of_month) + '日')
                    continue
                if list(row)[10] == '缺卡':
                    if query_data(name_names, '上班未打卡') == '':
                        update_data(name_names, '上班未打卡',
                                    query_data(name_names, '上班未打卡') + str(day_of_month) + '日')

                    else:
                        update_data(name_names, '上班未打卡',
                                    query_data(name_names, '上班未打卡') + '、' + str(day_of_month) + '日')
                    if query_data(name_names, '旷工') == '':
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + str(day_of_month) + '日上午')
                    else:
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + '、' + str(day_of_month) + '日上午')
                if list(row)[16] == '缺卡':

                    if query_data(name_names, '下班未打卡') == '':
                        update_data(name_names, '下班未打卡', str(day_of_month) + '日')
                    else:
                        update_data(name_names, '下班未打卡',
                                    query_data(name_names, '下班未打卡') + '、' + str(day_of_month) + '日')
                    if query_data(name_names, '旷工') == '':
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + str(day_of_month) + '日下午')
                    else:
                        update_data(name_names, '旷工',
                                    query_data(name_names, '旷工') + '、' + str(day_of_month) + '日下午')
                if list(row)[16] == '正常' or list(row)[16] == '早退':
                    update_data(name_names,time_category(list(row)[15]),query_data(name_names, time_category(list(row)[15]))+1)
                if pd.isna(row.iloc[16]):
                    # print(str(row.iloc[12]))
                    if str(row.iloc[12]) == '正常' or '补卡审批通过' in str(row.iloc[12]) or list(row)[16] == '早退':
                        update_data(name_names, time_category(list(row)[11]),
                                    query_data(name_names, time_category(list(row)[11])) + 1)
                        if str(row.iloc[10]) == '正常':
                            update_data(name_names, '打卡', query_data(name_names, '打卡') + 1)
                        if list(row)[8] == '休息':
                            update_data(name_names, '加班', query_data(name_names, '加班') + 1)
                            if query_data(name_names, '非展会周末加班') == '':
                                update_data(name_names, '非展会周末加班', str(day_of_month) + '日加班' + str(
                                    time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
                            else:
                                update_data(name_names, '非展会周末加班',
                                            query_data(name_names, '非展会周末加班') + '、' + str(
                                                day_of_month) + '日加班' + str(
                                                time_difference_jiaban(list(row)[9], list(row)[11])) + '天')
print(json_data)
print(names)
# input()
wb2 = load_workbook(filename=file_name2)

# 获取工作表
sheet = wb2.worksheets[0]
kk = 0
# 临时变量名字
name_name = ''

for row in sheet.rows:
    # 遍历每一列
    kk += 1
    if kk <= 3:
        continue
    list_data1 = []
    for cell in row:
        # 打印单元格的值
        list_data1.append(cell.value)
    if list_data1[2] != None:
        name_name = list_data1[2]
        leave_day = query_data(name_name, '调休请假') if query_data(name_name, '调休请假') is not None else 0
        print(work_day, leave_day,query_data(name_name, '加班'))
        sheet.cell(row=kk, column=4).value = work_day - leave_day + query_data(name_name, '加班')
        sheet.cell(row=kk, column=9).value = query_data(name_name, '请假')
        sheet.cell(row=kk, column=5).value = query_data(name_name, '调休请假')
        sheet.cell(row=kk, column=8).value = query_data(name_name, '非展会周末加班')
    # print(kk, name_name)
    if list_data1[5] == '加班到10点以后':
        sheet.cell(row=kk, column=7).value = query_data(name_name, '加班到10点以后')
    if list_data1[5] == '加班到9点以后':
        sheet.cell(row=kk, column=7).value = query_data(name_name, '加班到9点以后')
    if list_data1[5] == '加班到8点以后':
        sheet.cell(row=kk, column=7).value = query_data(name_name, '加班到8点以后')
    if list_data1[5] == '7点以后下班':
        sheet.cell(row=kk, column=7).value = query_data(name_name, '7点以后下班')
    if list_data1[5] == '7点以前下班':
        sheet.cell(row=kk, column=7).value = query_data(name_name, '7点以前下班')

# 保存工作簿
wb2.save(filename=file_name2)


















# 加载工作簿
wb1 = load_workbook(filename=file_name)

# 获取工作表
sheet = wb1.worksheets[1]
kk = 0
for row in sheet.rows:
    # 遍历每一列
    kk += 1
    if kk <= 2:
        continue
    list_data1 = []
    for cell in row:
        # 打印单元格的值
        list_data1.append(cell.value)
    # 上班未打卡列
    sheet.cell(row=kk, column=5).value = query_data(list_data1[3], '上班未打卡')
    # 下班未打卡列
    sheet.cell(row=kk, column=8).value = query_data(list_data1[3], '下班未打卡')
    # 迟到20分钟内列
    sheet.cell(row=kk, column=6).value = query_data(list_data1[3], '迟到20分钟内')
    # 迟到20分钟以上列
    sheet.cell(row=kk, column=7).value = query_data(list_data1[3], '迟到20分钟以上')
    # 旷工列
    sheet.cell(row=kk, column=9).value = query_data(list_data1[3], '旷工')
    #调休
    sheet.cell(row=kk, column=10).value = query_data(list_data1[3], '调休')
    # 请假列
    sheet.cell(row=kk, column=11).value = query_data(list_data1[3], '请假')
    # 外出 如果query_data(list_data1[3], '外出')是0则不显示
    if query_data(list_data1[3], '外出') == 0:
        sheet.cell(row=kk, column=14).value = ''
    else:
        sheet.cell(row=kk, column=14).value = query_data(list_data1[3], '外出')
# 修改指定位置的数据
# sheet.cell(row=13, column=13).value = '123'

# 保存工作簿
wb1.save(filename=file_name)



